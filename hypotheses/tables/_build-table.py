#!/usr/bin/env python3
"""
hypotheses/tables/_build-table.py

Rebuild hypotheses.xlsx + hypotheses.csv from MD files в hypotheses/{active,testing,confirmed,refuted,paused,samples}/
+ rebuild alphas-state-graph.xlsx from per-hypothesis alphas frontmatter + alpha-trail files.

Usage:
  cd ~/jetix-os
  python3 hypotheses/tables/_build-table.py [--csv-only] [--no-include-samples] [--no-alphas-table]

Layer 7 of 7-layer hypotheses architecture.
Filesystem-authoritative: MD primary; Excel/CSV = derived view.
Safe re-run (idempotent).
"""
import os
import sys
import argparse
from pathlib import Path

try:
    import yaml
except ImportError:
    print("FAIL: PyYAML required (pip install pyyaml)", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

REPO_ROOT = Path(__file__).resolve().parents[2]
HYPOTHESES_DIR = REPO_ROOT / "hypotheses"
STATUS_DIRS = ["active", "testing", "confirmed", "refuted", "paused"]
SAMPLES_DIR = "samples"
TABLES_DIR = HYPOTHESES_DIR / "tables"


def parse_frontmatter(md_path):
    """Extract YAML frontmatter from a markdown file. Returns dict or None."""
    try:
        with open(md_path, encoding="utf-8") as f:
            content = f.read()
    except Exception as e:
        print(f"WARN: cannot read {md_path}: {e}", file=sys.stderr)
        return None
    if not content.startswith("---"):
        return None
    # Find closing ---
    end = content.find("\n---", 4)
    if end == -1:
        return None
    yaml_block = content[4:end]
    try:
        return yaml.safe_load(yaml_block)
    except yaml.YAMLError as e:
        print(f"WARN: yaml parse error в {md_path}: {e}", file=sys.stderr)
        return None


def collect_hypotheses(include_samples=True):
    """Walk все status dirs + samples; collect hypothesis frontmatter."""
    rows = []
    dirs = list(STATUS_DIRS)
    if include_samples:
        dirs.append(SAMPLES_DIR)
    for d in dirs:
        path = HYPOTHESES_DIR / d
        if not path.exists():
            continue
        for md_file in sorted(path.glob("H-*.md")):
            fm = parse_frontmatter(md_file)
            if fm and isinstance(fm, dict):
                fm = dict(fm)  # shallow copy
                fm["_file"] = str(md_file.relative_to(REPO_ROOT))
                fm["_status_dir"] = d
                rows.append(fm)
    return rows


def _flatten(val):
    """Convert lists/dicts к string для tabular display."""
    if val is None:
        return ""
    if isinstance(val, (list, dict)):
        try:
            return yaml.safe_dump(val, allow_unicode=True, default_flow_style=True).strip()
        except Exception:
            return str(val)
    return val


def write_csv(rows, csv_path):
    """Write CSV. Pandas-preferred; fallback к csv module."""
    if not rows:
        # Write empty file with header
        import csv as _csv
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            f.write("id,title,status,category,lifecycle\n")
        return

    fieldnames = sorted({k for row in rows for k in row.keys()})
    # Put canonical fields first
    canonical = ["id", "title", "slug", "status", "_status_dir", "category", "lifecycle",
                 "owner", "created", "updated", "fpf_F", "fpf_G", "fpf_R",
                 "strategic_region", "test_method", "success_criteria", "refutation_criteria",
                 "outcome", "learning_extracted", "alphas", "linked_artefacts", "linked_hypotheses",
                 "_file"]
    ordered = [f for f in canonical if f in fieldnames] + [f for f in fieldnames if f not in canonical]

    if HAS_PANDAS:
        flat_rows = [{k: _flatten(v) for k, v in r.items()} for r in rows]
        df = pd.DataFrame(flat_rows)
        df = df.reindex(columns=ordered)
        df.to_csv(csv_path, index=False, encoding="utf-8")
    else:
        import csv as _csv
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            w = _csv.DictWriter(f, fieldnames=ordered, extrasaction="ignore")
            w.writeheader()
            for row in rows:
                w.writerow({k: _flatten(v) for k, v in row.items() if k in ordered})


def write_xlsx(rows, xlsx_path):
    """Write Excel with bold header + autowidth."""
    if not HAS_OPENPYXL:
        print("openpyxl not available — skipping xlsx", file=sys.stderr)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hypotheses"

    if not rows:
        ws.cell(row=1, column=1, value="No hypotheses yet")
        wb.save(xlsx_path)
        return

    fieldnames = sorted({k for row in rows for k in row.keys()})
    canonical = ["id", "title", "slug", "status", "_status_dir", "category", "lifecycle",
                 "owner", "created", "updated", "fpf_F", "fpf_G", "fpf_R",
                 "strategic_region", "test_method", "success_criteria", "refutation_criteria",
                 "outcome", "learning_extracted", "alphas", "linked_artefacts", "linked_hypotheses",
                 "_file"]
    ordered = [f for f in canonical if f in fieldnames] + [f for f in fieldnames if f not in canonical]

    # Header
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)
    for col_idx, name in enumerate(ordered, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, name in enumerate(ordered, 1):
            val = _flatten(row.get(name, ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Auto-width (rough)
    for col_idx, _ in enumerate(ordered, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        # Heuristic: cap к 50 chars
        max_len = max(
            len(str(_flatten(r.get(ordered[col_idx - 1], "")))[:80]) for r in rows
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)

    ws.freeze_panes = "A2"
    wb.save(xlsx_path)


def write_alphas_xlsx(rows, xlsx_path):
    """Write alphas-state-graph.xlsx — sheet per alpha + per-hypothesis state."""
    if not HAS_OPENPYXL:
        print("openpyxl not available — skipping alphas-state-graph.xlsx", file=sys.stderr)
        return

    ALPHAS = ["stakeholder", "opportunity", "requirements", "software-system",
              "work", "team", "way-of-working"]

    wb = openpyxl.Workbook()
    # Remove default sheet, will re-add per alpha
    wb.remove(wb.active)

    header_fill = PatternFill("solid", fgColor="D6F0D6")
    header_font = Font(bold=True)

    for alpha_name in ALPHAS:
        ws = wb.create_sheet(title=alpha_name[:31])  # Excel max title 31 chars
        # Header row
        header = ["H-NNN", "Hypothesis title", "Current state", "Status (hypothesis)",
                  "File", "Last updated"]
        for col_idx, name in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill

        # Find rows where alphas list contains this alpha
        row_idx = 2
        for r in rows:
            alphas_list = r.get("alphas", []) or []
            if not isinstance(alphas_list, list):
                continue
            for alpha_entry in alphas_list:
                if not isinstance(alpha_entry, dict):
                    continue
                if alpha_entry.get("alpha") == alpha_name:
                    ws.cell(row=row_idx, column=1, value=r.get("id", ""))
                    ws.cell(row=row_idx, column=2, value=r.get("title", ""))
                    ws.cell(row=row_idx, column=3, value=alpha_entry.get("state", ""))
                    ws.cell(row=row_idx, column=4, value=r.get("status", ""))
                    ws.cell(row=row_idx, column=5, value=r.get("_file", ""))
                    ws.cell(row=row_idx, column=6, value=str(r.get("updated", "")))
                    row_idx += 1
        # Auto-width
        for col_letter, width in zip(["A", "B", "C", "D", "E", "F"], [10, 50, 25, 15, 50, 12]):
            ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = "A2"

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description="Rebuild hypotheses Excel/CSV tables (Layer 7)")
    parser.add_argument("--csv-only", action="store_true", help="Skip xlsx output")
    parser.add_argument("--no-include-samples", action="store_true", help="Exclude samples/")
    parser.add_argument("--no-alphas-table", action="store_true", help="Skip alphas-state-graph.xlsx")
    args = parser.parse_args()

    include_samples = not args.no_include_samples

    rows = collect_hypotheses(include_samples=include_samples)
    print(f"Collected {len(rows)} hypotheses (include_samples={include_samples})")

    TABLES_DIR.mkdir(exist_ok=True)

    csv_path = TABLES_DIR / "hypotheses.csv"
    write_csv(rows, csv_path)
    print(f"Wrote {csv_path}")

    if not args.csv_only:
        xlsx_path = TABLES_DIR / "hypotheses.xlsx"
        write_xlsx(rows, xlsx_path)
        print(f"Wrote {xlsx_path}")

        if not args.no_alphas_table:
            alphas_xlsx = TABLES_DIR / "alphas-state-graph.xlsx"
            write_alphas_xlsx(rows, alphas_xlsx)
            print(f"Wrote {alphas_xlsx}")

    print(f"Done. {len(rows)} rows.")


if __name__ == "__main__":
    main()
